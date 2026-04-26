import asyncio
import os
import subprocess
import io
import re
import requests

from datetime import datetime, timedelta
from copy import copy

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side
from pdf2image import convert_from_path
import xlrd

from app.core.config import settings
from app.core.s3 import s3
from app.router.group_router import send_group


class ParsAask:
    def __init__(self):
        self.GROUP_NAMES = []

    async def download_and_generate_schedule(self, manual_url: str = None):
        today = datetime.now()
        if today.weekday() == 6:
            today += timedelta(days=1)

        i = 0

        while True:
            try:
                target_day = today + timedelta(days=i)
                day_month = int(target_day.strftime("%d%m"))

                if manual_url:
                    url = manual_url
                    manual_url = None
                else:
                    url = f"https://altask.ru/images/raspisanie/DO/{day_month}.xls"

                response = requests.get(url)

                file_path = f"{day_month}.xls"
                with open(file_path, "wb") as f:
                    f.write(response.content)

                await self.extract_group_names_from_xls(file_path)

                self.parse_and_generate_tables(file_path, day_month)

                i += 1

            except Exception:
                groups_ary = list(set(self.GROUP_NAMES))
                send_group(groups_ary, address_name="пр.Ленина 68")
                break

    async def extract_group_names_from_xls(self, file_path):
        group_pattern = re.compile(r"^[А-Яа-яA-Za-z]+-\d{2}$")
        group_names = set()

        workbook = xlrd.open_workbook(file_path)
        sheet = workbook.sheet_by_index(0)

        for y in range(sheet.nrows):
            for x in range(sheet.ncols):
                value = str(sheet.cell_value(y, x)).strip()
                if group_pattern.match(value):
                    group_names.add(value)

        self.GROUP_NAMES = sorted(group_names, key=lambda x: x.lower())


        print(f"[INFO] Найдено групп: {len(self.GROUP_NAMES)}")

    @staticmethod
    def convert_xls_to_xlsx(input_path):
        try:
            subprocess.run(
                [
                    "libreoffice",
                    "--headless",
                    "--convert-to",
                    "xlsx",
                    input_path,
                ],
                check=True,
            )
            return True
        except subprocess.CalledProcessError as e:
            print(f"[!] Ошибка конвертации: {e}")
            return False

    def read_xls_file(self, file_path):
        result = []

        workbook = xlrd.open_workbook(file_path)
        sheet = workbook.sheet_by_index(0)

        for y in range(sheet.nrows):
            for x in range(sheet.ncols):
                value = sheet.cell_value(y, x)

                if value in self.GROUP_NAMES:
                    ly = y + 1

                    while (
                        ly - y < 13
                        and ly < sheet.nrows
                        and sheet.cell_value(ly, x) not in self.GROUP_NAMES
                    ):
                        ly += 1

                    if ly >= sheet.nrows:
                        ly = sheet.nrows - 1

                    result.append(
                        {"group": value, "x": x, "y1": y + 1, "y2": ly}
                    )

        return result

    @staticmethod
    def upload_image_to_s3(image, day_month, group_name):
        buffer = io.BytesIO()
        image.save(buffer, format="PNG")
        buffer.seek(0)

        s3_key = f"ААСК/пр.Ленина 68/{day_month}/{group_name}.png"

        s3.put_object(
            Bucket=settings.S3_BUCKET,
            Key=s3_key,
            Body=buffer,
            ContentType="image/png",
        )

        print(f"[S3] Загружено: {s3_key}")

        return s3_key

    def create_group_sheets_single_column(self, groups, source_sheet, day_month):
        for group in groups:
            x = group["x"] + 1
            y1 = group["y1"]
            y2 = group["y2"]
            name = group["group"]

            wb = Workbook()
            ws = wb.active
            ws.title = name

            for row in range(y1, y2 + 1):
                src_cell = source_sheet.cell(row=row, column=x)
                tgt_cell = ws.cell(row=row - y1 + 1, column=4, value=src_cell.value)

                if src_cell.has_style:
                    tgt_cell.font = copy(src_cell.font)

                    border = Border(
                        left=Side(style="medium"),
                        right=Side(style="medium"),
                        top=src_cell.border.top or Side(style="thin"),
                        bottom=src_cell.border.bottom or Side(style="thin"),
                    )

                    tgt_cell.border = border
                    tgt_cell.fill = copy(src_cell.fill)
                    tgt_cell.alignment = copy(src_cell.alignment)

            col_letter_src = source_sheet.cell(row=y1, column=x).column_letter
            src_width = source_sheet.column_dimensions[col_letter_src].width

            ws.column_dimensions["D"].width = max(
                src_width if src_width else 20,
                33,
            )

            for row in range(y1, y2 + 1):
                if source_sheet.row_dimensions[row].height is not None:
                    ws.row_dimensions[row - y1 + 1].height = (
                        source_sheet.row_dimensions[row].height
                    )

            temp_xlsx = f"{name}.xlsx"
            temp_pdf = f"{name}.pdf"

            wb.save(temp_xlsx)

            try:
                subprocess.run(
                    [
                        "libreoffice",
                        "--headless",
                        "--convert-to",
                        "pdf",
                        temp_xlsx,
                    ],
                    check=True,
                )

                images = convert_from_path(temp_pdf, dpi=300)

                if images:
                    img = images[0]
                    width, height = img.size

                    cropped = img.crop(
                        (
                            int(width * 0.33),               # left: отступ слева от ширины
                            0,                               # upper: начинаем с самого верха (0px)
                            width - int(width * 0.15),        # right: обрезаем справа % ширины
                            height - int(height * 0.10),     # lower: обрезаем снизу % высоты
                        )
                    )

                    self.upload_image_to_s3(cropped, day_month, name)

            except Exception as e:
                print(f"[!] Ошибка для группы {name}: {e}")

            finally:
                if os.path.exists(temp_xlsx):
                    os.remove(temp_xlsx)
                if os.path.exists(temp_pdf):
                    os.remove(temp_pdf)

    def parse_and_generate_tables(self, input_xls, day_month):
        self.convert_xls_to_xlsx(input_xls)

        groups = self.read_xls_file(input_xls)

        workbook = load_workbook(f"{input_xls}x")
        sheet = workbook.active

        self.create_group_sheets_single_column(groups, sheet, day_month)

        if os.path.exists(input_xls):
            os.remove(input_xls)

        if os.path.exists(f"{input_xls}x"):
            os.remove(f"{input_xls}x")

        print("[INFO] Обработка завершена")

        return True


parse_aask = ParsAask()